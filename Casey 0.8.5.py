#Casey
#Copyright 2017
#All rights arising under Copyright Act 1968 (Cth) are reserved 
#The moral rights of the author are asserted

import traceback
import sys
import requests
import bs4
import re
import datetime
import string
import csv

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

import statistics
from collections import Counter
import numpy
#import nltk.sentiment

#maybe move some of these imports to later.

#Introducing variables
jur = "nil"
court = "nil"
case_number = 0
appeal_allowed_count = 0
appeal_denied_count = 0
other_count = 0
affirm_citation_list = []
allowed_citation_list = []
case_dictlist = list()

n = -1

#DEFINING FUNCTIONS
#----------------------------
def super_regex(pattern, *args):
    for location in args:
        result = re.search(pattern, location)
        if result is not None:
            return result    

def remove_statute_section_fullstops(text):
    text = re.sub("(?<=s)\.(?=\d)", " ", text)
    return text

def excel_export(dicto):
    for bit in range(0 , len(case_dict[str(title)])):
        titlecell = ws1.cell( row = 3 + len(case_dictlist) , column = 2)
        titlecell.value = str(title)         
        cell = ws1.cell( row = 3 + len(case_dictlist) , column = 3+bit)
        cell.value = case_dict[str(title)][bit]

def neg_markup(textlist):
    textstring = " ".join(textlist)
    textstring = remove_statute_section_fullstops(textstring)
    textstring = re.sub("\.\s", "\s\.\s", textstring)
    textlist = textstring.split()

    marked_text = nltk.sentiment.util.mark_negation(str(textlist).split(), False, False)
    marked_text= " ".join(marked_text)
    return marked_text

#Regex artifact removal function
def remove_brackets_quotes(infected_string):
    try:
        infected_string.replace("[","").replace("]","").replace("'","").replace("(","").replace(")","").replace("(","").replace(": ","")
    except AttributeError:
        infected_string = re.sub("\[" , "",str(infected_string))
    return str(infected_string)

#Getting date object
def get_date(date_regex):
    try:
        date_regex = date_regex.group()
    except AttributeError:
        pass

    date = str(date_regex)
    date = str(date_regex).replace("[","").replace("]","").replace("'","").replace("(","").replace(")","").replace("(","").replace(": ","")
    date = re.sub(r'^[^\d]*', '', date)
    date.strip()
    date = re.sub("\d\d(?=\d\d)", "", str(date)) #get rid of first two numbers (e.g. 20) in the year
    date = re.sub("\((?=\d\s)", "0", str(date)) #padding
    if "fedcourt" in sauce:
        date = re.sub('(?<=d)\s(?=\d)', '', date)
    
    if date == "heard on the papers":
        date == "heard on the papers"
    else:
        try:
            date = datetime.datetime.strptime(str(date), "%d %B %y")
        
        except ValueError:
            date = "N/A"
    return date

#Strip Judge Names function
def strip_judge_name(judge):
    judge
    judge = re.sub('\n', ' ', judge)
    judge = re.sub('\t', ' ', judge)
    judge = re.sub(',', ' ', judge)
    judge = re.sub('\.','  ', judge) #to keep spacing. single space replace not working
    judge = re.sub('member', ' ', judge)
    judge = re.sub('senior', ' ', judge)
    judge = re.sub('deputy', ' ', judge)
    judge = re.sub('president', ' ', judge)
    judge = re.sub('tribunal', ' ', judge)
    judge = re.sub(':', ' ', judge)
    judge = re.sub(' ao', ' ', judge)
    judge = re.sub('dr ', ' ', judge)
    judge = re.sub(' rfd', ' ', judge)
    judge = re.sub('mrs ', ' ', judge)
    judge = re.sub('mr ', ' ', judge) #mr must be after mrs
    judge = re.sub('ms ', ' ', judge)
    judge = re.sub('vice ', ' ', judge)
    judge = re.sub('judge ', ' ', judge)
    judge = re.sub('judge(s)', ' ' , judge)
    judge = re.sub('  ', '', judge)
    judge = re.sub('\)', '', judge)
    judge = re.sub('\(', '', judge)
    judge = judge.strip()
    return judge
#brackets need to be taken out

#Soupify HTML
def simplify_html(html):
    #one = all text, no tags
    #two = no excessive whitespace
    #three = lowercase
    #four = no whitespace characters anymore
    text_one = html.get_text(" ", strip=True)
    text_two = re.sub("  ", "", text_one)
    text_three = text_two.lower()
    text_four = text_three.replace('\n', ' ').replace('\t', ' ')
    return text_one, text_two, text_three, text_four

#Aggregrate a variable based on case success/fail into Counter objects
def aggregrate_success_fail(success_list, fail_list):
    success_collection = Counter(success_list)
    fail_collection = Counter(fail_list)
    combined_collection = fail_collection + success_collection
    return fail_collection, success_collection, combined_collection

#Getting the Success rate via list comprehension of Counter objects
def get_success_rate(fail_collection, combined_collection):
    a = fail_collection
    b = combined_collection
    c = Counter({k:(b[k]-a[k])/b[k]*100 for k in b})
    return c

#troglodyte maths
def mean_stdev_calculation(combined_collection, affirm_collection, success_rate):
    weighted_rates = Counter({k:(success_rate[k]*combined_collection[k]) for k in combined_collection})
    weighted_rate_mean = statistics.mean(weighted_rates.values())
    weighted_stdev = statistics.stdev(weighted_rates.values())

    unweighted_rate_mean = statistics.mean(success_rate.values())
    unweighted_stdev = statistics.stdev(success_rate.values())

    return weighted_rates, weighted_rate_mean, unweighted_rate_mean, unweighted_stdev, weighted_stdev

#filter1 = e.g. migration, *args is stuff to isolate
def single_filter(filter1, filter1_dict_position, *args):
    for x in case_dictlist[x].values():
        if case_dictlist[x].values()[filter1_dict_position] == filter1:
            print(filter1)
            #to be continued


#Court selection
court = input("""What institution are you interested in today? The following institutions are supported, please enter the code for the desired institution:

Administrative Appeals Tribunal = AATA
Fair Work Commission = FWC
Federal Court of Australia = FCA
Federal Circuit Court of Australia = FCCA

""")

if court == "AATA" or "FWC" or "FCCA" or "FCA":
    jur = "cth"
print("You chose" , court)

#Target selection
target = input("""Choose one of the codes from the square brackets below (case sensitive)

ADMINISTRATIVE APPEALS TRIBUNAL
-----------------------------------
All Appeals                   [AAO]    
Migration                     [MIG]
Refugee                       [RFG]     
Freedom of Information        [FOI]    
   
FEDERAL CIRCUIT COURT OF AUSTRALIA
FEDERAL COURT OF AUSTRALIA
-----------------------------------
Judicial Review & Appeal      [ADJR]
Fair Work Ombudsman actions   [FWO]      

FAIR WORK COMMISSION
-----------------------------------
Unfair Dismissal & Costs      [UD&C] 

""")

#Initial File Open and Printing
#text logs
f = open(" %a  results.rtf" %court, "w", encoding="utf-8")
f_two = open(" %a  stats.rtf" %court, "w", encoding="utf-8")

#spreadsheet
wb = Workbook()
dest_filename = 'CaseyOutput.xlsx'
ws1 = wb.active
ws1.title = "Case Dicts"
ws1.cell(row = 3, column = 2).value = "Case Title"
ws1.cell(row = 3, column = 3).value = "Judge"
ws1.cell(row = 3, column = 4).value = "Case Matter"
ws1.cell(row = 3, column = 5).value = "Plaintiff fail"
ws1.cell(row = 3, column = 6).value = "Plaintiff success"
ws1.cell(row = 3, column = 7).value = "Citations"
ws1.cell(row = 3, column = 8).value = "Citation count"
ws1.cell(row = 3, column = 9).value = "Costs"

#ACCESSING HOST DATABASE
#Setting case & year range$
#Select case and year range here
case_range = range(100,200)
year_range = range(2017, 2018)

#Setting mode
#Do you want to import a case

#Beginning Variables & HTML manipulation
for year in year_range:
    for case_number in case_range:
        try:
            #Getting sauce
            #Notice: The Fair Work Commission asserts copyright in its material: https://www.fwc.gov.au/about-us/legal/copyright

            if court == "FCA":
                sauce = "http://www.judgments.fedcourt.gov.au/judgments/Judgments/fca/single/"+str(year)+"/"+str(year)+"fca"+str(str(case_number).zfill(4))
            if court == "FCCA":
                sauce = 
            if court == "FWC":
                sauce = "http://www.fwc.gov.au/documents/decisionssigned/html/"+str(year)+"fwc"+str(case_number)+".htm"
            if court == "AATA":
                sauce =

            headers = {'User-Agent': 'Safari/537.36'}
            r = requests.get(sauce , headers=headers)

            #Manipulating sauce
            soup = bs4.BeautifulSoup(r.content.decode('utf-8', 'replace') , "html.parser")
            souptext_one = simplify_html(soup)[0]
            souptext_two = simplify_html(soup)[1]
            souptext_three = simplify_html(soup)[2]
            souptext_four = simplify_html(soup)[3]
            #print(souptext_four)

            #Getting title
            title = re.findall('<title[^>]*>([^<]+)</title>', str(soup))
            remove_brackets_quotes(str(title))
            title = " ".join(title)
               
            #Check r + Print sauce
            if str(r) == "<Response [200]>":
                case_number = int(case_number) + 1
                print(sauce , r, str(title))
                
            elif str(r) == "<Response [404]>":
                case_number = int(case_number) + 1
                print("No case at this number")
                continue

            #Checking titles to discriminate beteen cases
            if target == "MIG":
                if str(title).find("(Migration)") == -1 and str(title).find("Minister for Immigration)") == -1:
                    print("not a migration case")
                    continue

            elif target == "FOI":
                if str(title).find("(Freedom of Information)") == -1 and str(title).find("(Freedom of information)") == -1:
                    print("not a FOI case")
                    continue
                
            elif target == "RFG":
                if str(title).find("(Refugee)") == -1:
                    print("not a refugee case")
                    continue

            elif target == "FWO":
                if str(title).find("Fair Work Ombudsman v") == -1:
                    if str(title).find("FWO v") == -1:
                        print("not a FWO action case")
                        continue
                
    #SCANNING FOR INFO
            if court == "FCA":
                #Top Info Section
                info_section = re.search("(?<=original word document).*?(?=reasons for judgment|introduction)", souptext_four)
                info_section = info_section.group()
                
                #Case Below
                any_case_below = re.search("(?<=appeal from:) .+? \[\d\d\d\d\] .+? \d+", info_section, re.DOTALL)
                if any_case_below:
                    any_case_below = any_case_below.group()
                    print(any_case_below)
                
                #Discriminating
                if target == "ADJR" and any_case_below == None:
                    continue
                if re.search("refugee review|migration review|minister for|administrative appeals", info_section) == None:
                    continue
    
                #Catchwords
                catchwords_section = re.search("(?<=catchwords:).*(?=legislation)", info_section)
                #if legislation as a word comes up you are in trouble thou

                if catchwords_section == None: catchwords = "no catchwords"
                else:
                    catchwords = catchwords_section.group()
                print(catchwords)

                #Case Matter
                case_matter = re.search("(?<=catchwords:).*?(?=-|–|—|–|−)", souptext_four)
                if case_matter == None:
                    case_matter = re.search("\([a-z\s]+\)", str(title))
                    if case_matter == None:
                        case_matter = re.search(" ", souptext_four)
                case_matter = case_matter.group().strip()
                print(case_matter)

                #Judge
                judge_string = re.search("(?<=judge:) \w+ \w+ (?=date)", info_section)
                if judge_string == None:
                    #judge_string = re.search("judge.*?(?=date)", info_section)
                    #judge_string = re.sub(" " , "", judge_string)
                    #judge_string = judge_string.replace("judge", "").replace("(s)", "")
                    #need another lookahead here to avoid getting stuff at top of bar.
                    judge_string = re.search(" ", info_section)
                judge_string = judge_string.group()
                strip_judge_name(judge_string)
                print(judge_string)

                #Registry
                registry = re.search("(?<=registry:).*?(?=division)", info_section)
                registry = registry.group()
                print(registry)

                #Division
                division = re.search("(?<=national practice area:).+?(?!\(npas\))(?=category)", info_section)
                if division == None:
                    division = "unknown"
                else:
                    division = division.group()

                #Cases cited
                info_section_cases = re.search("(?<=cases cited).+?(?=date of hearing)", info_section)
                if info_section_cases == None: info_section_cases = re.search(" ", info_section)
                citation_list = re.findall("\[\d\d\d\d\] [a-z]+? \d{1,4}\s;\s\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}|\[\d\d\d\d\] [a-z]+? \d{1,4}|\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}", str(info_section_cases.group()))
                print(citation_list)

                #Results
                if target == "ADJR":
                    affirm_search = re.search("(?<=date of orders)(application is dismissed)|(be dismissed)|(application dismissed)|(refused)|(appeal dismissed)|(appeal is dismissed)", info_section) 
                    allow_search = re.search("(?<=date of orders)(is allowed)|(appeal allowed)|(is varied)|(is quashed)|(quashes)|(is set aside)|(sets aside)|(remits)|( writ )|(mandamus)|(certiorari)", info_section)
                else:
                    affirm_search = re.search("(?<=date of orders)|(application is dismissed)|(application dismissed)|(be dismissed)", info_section) 
                    allow_search = re.search("(?<=date of orders)(the court declares)|(declaration)", info_section)
            
                #Costs
                costs_order = re.search("(?<=orders).*?(applicant(s)? pay.*?)|(respondent(s)? pay.*?)", info_section)
                if costs_order: 
                    costs_amount = re.search("\$[\d,]+", info_section)
                    if costs_amount: 
                        costs_order = costs_amount.group()
                        costs_order = int(costs_order.replace("," , "").replace("$" , "")) #removing commas from number to mutate properly
                    else:
                        costs_order = "Costs awarded Amount unknown"
                if costs_order == None:
                    costs_order = "No costs order"
                print(costs_order)

                    
            if court == "FCCA":
                #Top Info Section
                info_section = re.search(".*?(?=reasons for judgment|introduction)", souptext_four)
                info_section = info_section.group()
                
                #Case Below
                any_case_below = re.search("(?<=appeal from:) .*? \[\d\d\d\d\] .+? \d+", info_section, re.DOTALL)
                if any_case_below:
                    any_case_below = any_case_below.group()
                    print(any_case_below)
                
                #Catchwords
                catchwords_section = re.search("(?<=catchwords).*?(?=legislation)", souptext_four)
                #if legislation as a word comes up you are in trouble thou

                if catchwords_section == None: catchwords = "no catchwords"
                else:
                    catchwords = catchwords_section.group()

                #Case Matter
                case_matter = re.search("(?<=catchwords:).*?(?=-|–|—|–|−)", souptext_four)
                if case_matter == None:
                    case_matter = re.search("\([a-z\s]+\)", str(title))
                    if case_matter == None:
                        case_matter = re.search(" ", souptext_four)
                case_matter = case_matter.group().strip()

                #Judge
                judge_string = re.search("(?<=judgment of:).*?(?=hearing dat)", info_section)
                judge_string = judge_string.group()
                strip_judge_name(judge_string)
                print(judge_string)

                #Registry
                registry = re.search("(?<=delivered at:) \w+", souptext_four)
                registry = registry.group()
                print(registry)

                #Division
                division = "not bothering"

                #Cases cited
                citation_list = re.findall("\[\d\d\d\d\] [a-z]+? \d{1,4}\s;\s\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}|\[\d\d\d\d\] [a-z]+? \d{1,4}|\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}", info_section)
                print(citation_list)

                #Results
                if target == "ADJR":
                    if re.search("and minister|respondent: administrative appeals tribunal|respondent: minister|respondent: refugee review|respondent: migration review", info_section) == None:
                        continue
                    affirm_search = re.search("(?<=orders).*?(application is dismissed)|(application dismissed)|(be dismissed)|(no jurisdictional error).*?(?=federal circuit court)", info_section)
                    allow_search = re.search("(?<=orders).*?(is allowed)|(appeal allowed)|(is varied)|(is quashed)|(quashes)|(is set aside)|(sets aside)|(remit)|( writ )|(mandamus)|(certiorari).*?(?=federal circuit court)", info_section)
                    if allow_search == None:
                        allow_search = re.search("((the court declares)|(declaration))(?=federal circuit court)", info_section)
                else:
                    affirm_search = re.search("(?<=date of orders)|(application is dismissed)|(application is dismissed)|(be dismissed)", info_section) 
                    allow_search = re.search("(?<=date of orders)(the court declares)|(declaration)", info_section)


                #Costs
                costs_order = re.search("(?<=orders).*?(applicant(s)? pay.*?)|(respondent(s)? pay.*?)", info_section)
                if costs_order: 
                    costs_amount = re.search("\$[\d,]+", info_section)
                    if costs_amount: 
                        costs_order = costs_amount.group()
                        costs_order = int(costs_order.replace("," , "").replace("$" , ""))
                    else:
                        costs_order = "Costs awarded Amount unknown"
                if costs_order == None:
                    costs_order = "No costs order"
                print(costs_order)

                
            if court == "FWC":
                #Top Info Section
                info_section = re.search("(decision|fair work commission).+?\[1\]", souptext_four)
                info_section = info_section.group()
                
                ##Catchwords
                #catchwords_section = re.search("\d+ \w+ \d\d\d\d .*", info_section)
                #if catchwords_section == None: catchwords = "no catchwords"
            
                #else:
                #    catchwords_section = catchwords_section.group()
                #    catchwords = catchwords_section
                #print(catchwords)
                catchwords = info_section
                
                #Case Matter
                case_matter = re.findall("(?<=fair work act 2009).*?s\.\d+", info_section)
                case_matter = " ".join(case_matter)
                case_matter = str(case_matter)      
                if case_matter == None:
                    case_matter = re.search(" ", souptext_four)
                    case_matter = case_matter.group()
                print(case_matter)
                print(type(case_matter))

                #Discriminating
                if str(case_matter).find("394") == -1 and str(case_matter).find("365") == -1 and str(case_matter).find("costs") == -1: 
                    continue
                
                #Division
                division = "no divisions in FWC"
                
                #Judge
                judge_string = re.search("(?<=\d\) )((?<=industry)?|(?<=operations)?)(commissioner .+?\s|(senior )?deputy president .+?\s|(vice )?president .+?\s)", info_section)
                judge_string = judge_string.group()
                strip_judge_name(judge_string)
                print(judge_string)

                #Registry
                registry = re.search("(?<={judge}).*?(?=,)".format(judge=judge_string), souptext_four)
                registry = registry.group()
                print(registry)

                #Paragraphs.
                all_paras = re.search("\[1\].*(?=var _)", souptext_four)
                first_para = re.search("\[1\].*(?=\[2\])", souptext_four)
                concluding_paras = re.search("((?<=[\.\"] conclusion \[\d).*)|((?<=[\.\"] conclusion and orders \[\d).*)", souptext_four)
                if concluding_paras == None:
                    concluding_paras = "nothing"
                else: concluding_paras = concluding_paras.group()

                concluding_paras_sentences = str(concluding_paras)
                concluding_paras_sentences = concluding_paras_sentences.split(". ")

                #Cased Cited
                info_section_cases = re.search("var _gaq.* ", souptext_four)
                if info_section_cases == None: info_section_cases = re.search(" ", souptext_four)
                citation_list = re.findall("\[\d\d\d\d\] [a-z]+? \d{1,4}\s;\s\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}|\[\d\d\d\d\] [a-z]+? \d{1,4}|\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}", str(info_section_cases.group()))   
                
                #Costs detection
                fwc_costs_detect = super_regex("(s.400A)|(s.611)|(s.401)|(costs)", str(catchwords), str(concluding_paras), str(info_section))
                if fwc_costs_detect is not None:
                    case_matter = "costs"
                    print(case_matter)
                costs_order = "no costs awarded"

                marked_concluding_paras = concluding_paras
                marked_catchwords = catchwords

                if target == "UD&C":
                    if case_matter == "costs":
                        affirm_search = super_regex("(is dismissed)|(be dismissed)|(application dismissed)|(refuse(d)?)|(611_NEG)||(400A_NEG)|(jurisdiction_NEG)|(vexatious_NEG)|(unreasonable_NEG)",
                                                    concluding_paras, marked_concluding_paras, catchwords, marked_catchwords)
                        
                        allow_search = super_regex("(award)|(indemnity)",
                                                catchwords, marked_catchwords)

                        if allow_search:
                            costs_order = super_regex("pay.*?\$[\d,]", marked_concluding_paras, catchwords, marked_catchwords)
                            costs_order = costs_order.group()
                            print(costs_order)

                    if "s.394" in case_matter:
                        allow_search = super_regex(
                            "[,\"] remedy \[\d+] \w", souptext_four)
                        
                        if allow_search == None:
                            allow_search = super_regex(
                                "(compensation.*?\$)|(compensation.*?week)|(reinstate)|(s.390)|(s.392)", concluding_paras, catchwords)
                        
                        affirm_search = super_regex(
                            "(does not satisfy)|(application is dismissed)|(application dismissed)|(not harsh,)|(no dismissal)|(not unfair)|(not dismissed)|(remedy is dismissed)|(must be dismissed)|(unfair_NEG)|(harsh_NEG)|(394_NEG)|(387_NEG)", marked_concluding_paras, catchwords, marked_catchwords)
                    
                    else:
                        allow_search = super_regex("placeholder", marked_concluding_paras, catchwords, marked_catchwords)
                        affirm_search = super_regex("placeholder", marked_concluding_paras, catchwords, marked_catchwords)
                

            if court == 'AATA':
                #Top Info Section
                info_section = re.search("(division).+?(catchwords|reasons for decision|legislation|sgd|\[1\]?)", souptext_four)
                info_section = info_section.group()

                #Costs
                costs_order = "no costs in AATA"

                #Division
                division = re.search("(?<=division)(?<=:)?.+?(?=division)", info_section)
                if division == None:
                    division = "unknown"
                else:
                    division = division.group()
                    division = division.replace(":","")
                    division = division.strip()
                
                #Catchwords
                catchwords_section = re.search("(?<=catchwords).*?(?=legislation)", souptext_four)
                #if legislation as a word comes up you are in trouble thou
                if catchwords_section == None: catchwords = "no catchwords"
                else:
                    catchwords = catchwords_section.group()
                
                #Case Matter —
                #gets the first catchword segment
                case_matter = re.search("(?<=catchwords).*?(?=-|–|—|–|−)", souptext_four)
                if case_matter == None:
                    case_matter = re.search("\([a-z\s]+\)", str(title))
                elif case_matter == None:
                    case_matter = re.search(" ", souptext_four)
                case_matter = case_matter.group()
                case_matter = case_matter.strip().replace(")" , "").replace("(" , "")

                #Cases Cited
                info_section_cases = re.search(
                    "cases .*? (secondary materials|reasons for decision|statement of decision)", souptext_four)
                if info_section_cases == None: info_section_cases = re.search(" ", souptext_four)
                citation_list = re.findall(
                    "\[\d\d\d\d\] [a-z]+? \d{1,4}\s;\s\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}|\[\d\d\d\d\] [a-z]+? \d{1,4}|\(\d\d\d\d\) \d{1,4} [a-z]+? \d{1,4}", str(info_section_cases.group()))
                print(citation_list)

                #Judge
                judge_string = re.search("((?<=tribunal:)|(?<=member:)|(?<=members:)|(?<=tribunal)|(?<=member)).*?((?=date)|(?=,))", info_section)
                judge_string = judge_string.group()
                strip_judge_name(judge_string)
                print(judge_string)

                #Registry
                registry = re.search("(place of decision(:)? \w+)|(place(:)? \w+)", souptext_four)
                registry = registry.group()
                registry = registry.replace("place of decision" , "").replace("place" , "").replace(":" , "")

                print(registry)

                affirm_search = re.search("(affirm the decision)|(is refused)|(tribunal affirms)|(is affirmed)|(tribunal dismisses)|(is dismissed)", info_section)
                allow_search = re.search("(is varied)|(tribunal varies)|(set aside)|(sets aside)|(tribunal remits)|(substitute)", info_section)

            #Avoiding judge NoneType error for all above
            if judge_string == None:
                judge_string = re.search(" ", str(soup))
                                    
    #SCANNING FOR RESULTS AND COUNTING
            if affirm_search:
                #Adding to appeal denied counter
                appeal_denied_count = appeal_denied_count + 1

                #Adding citations to affirm case list    
                for citation in citation_list:
                    affirm_citation_list.append(citation)

                #printing and writing results, with error handling
                print("Appeal denied" + judge_string)
                print(info_section)
                print(affirm_search.group())
                print("appeals denied =" , appeal_denied_count)
                f.write(" appeal denied " + str(title)+ judge_string + "\n")
            
            elif allow_search:                               
                #Adding to appeal allowed counter
                appeal_allowed_count = appeal_allowed_count + 1

                #Adding citations to affirm case list    
                for citation in citation_list:
                    allowed_citation_list.append(citation)

                #printing and writing results, with error handing
                print("Appeal allowed" + judge_string)
                print(info_section)
                print(allow_search.group())
                print("appeals allowed =" , appeal_allowed_count)
                f.write(" appeal allowed " + str(title) + judge_string + "\n")
                            
            #Unable to find the outcome
            else:
                print("Syntax not found. Please check manually " + judge_string)
                other_count = other_count + 1
                f.write(" Other " + str(title)+ judge_string + "\n")

    #CALCULATING TIMES
            #Collecting date data

            #Collecting Decision Date
            if "blah" in sauce:
                title_date_regex = re.findall("\(\d{1,2} \w+ \d\d\d\d", str(title))
                title_date = get_date(title_date_regex)
                #decision year
                
            if "fwc.gov" in sauce:
                title_date_regex = re.search("(?<=[a-z],).*?\d\d? \w+ \d\d\d\d", info_section)
                title_date_regex = title_date_regex.group()
                title_date = get_date(title_date_regex)

            if "fedcourt.gov" in sauce:
                title_date_regex = re.search("(?<=date of judgment:).*?\d\d\d\d", info_section)
                if title_date_regex == None:
                    title_date_regex = re.search("(?<=date of order:).*?\d\d\d\d", info_section)

                title_date_regex = title_date_regex.group()
                print(title_date_regex)
                title_date = get_date(title_date_regex)
                #, 4 JANUARY 2017

            if court == "FCA":
                #Regex for date data
                hearing_date_regex = re.findall("((?<=date of hearing)|(?<=date\(s\) of hearing)|(?<=dates of hearing)|(?<=of hearing)):?\d{1,2} \w+ \d\d\d\d", str(souptext_four))
                #put colon after every

                submission_date_regex = re.findall("(?<=date of last submission):?|(?<=date final submissions received):? \d{1,2} \w+ \d\d\d\d", str(souptext_four))

                #Normalise and convert date data to date object / End results
                hearing_date = get_date(hearing_date_regex)
                submission_date = get_date(submission_date_regex)
                title_date = get_date(title_date_regex)
                
            if court == "AATA":
                #Regex for date data
                hearing_date_regex = re.search("((?<=date of hearing)|(?<=date\(s\) of hearing)|(?<=dates of hearing)|(?<=of hearing)):? \d{1,2} \w+ \d\d\d\d", str(souptext_four))
                print(hearing_date_regex)
                submission_date_regex = re.search("((?<=date final submissions received)|(?<=date of last submission)):? \d{1,2} \w+ \d\d\d\d", str(souptext_four))
                if submission_date_regex == None: submission_date_regex = re.search(" " ,souptext_four)
                if hearing_date_regex == None: hearing_date_regex = re.search(" " ,souptext_four)
                
                #Normalise and convert date data to date object
                hearing_date = get_date(hearing_date_regex.group())
                submission_date = get_date(submission_date_regex.group())
                title_date = get_date(title_date_regex)

                print(title_date)
                print(hearing_date)
                print(submission_date)
            
            if court == "FCCA":
                #Regex for date data
                #actually needs to be hearing date: or hearing dates:
                hearing_date_regex = re.findall("(?<=date of hearing)|(?<=date\(s\) of hearing)|(?<=dates of hearing)|(?<=of hearing):?\d{1,2} \w+ \d\d\d\d", str(souptext_four))
                #put colon after every

                submission_date_regex = re.findall("(?<=date of last submission):?|(?<=date final submissions received):? \d{1,2} \w+ \d\d\d\d", str(souptext_four))

                #Normalise and convert date data to date object / End results
                hearing_date = get_date(hearing_date_regex)
                submission_date = get_date(submission_date_regex)
                title_date = get_date(title_date_regex)

            if court == "FWC":
                #Hearing Date
                #measures from first hearing
                
                if "hearing details:" not in souptext_four: hearing_date = "no hearing"
                else:
                    #Creating hearing soup
                    fwc_hearing_soup = re.search("(?<=hearing details:).*", souptext_four)
                    fwc_hearing_soup = fwc_hearing_soup.group()

                    #Getting hearing year
                    fwc_hearing_year = re.search("\d\d\d\d", fwc_hearing_soup)
                    fwc_hearing_year = fwc_hearing_year.group()


                    #Getting day-month
                    fwc_hearing_day_month = re.search("(\d\d? (january|february|march|april|may|june|july|august|september|october|november|december))|((january|february|march|april|may|june|july|august|september|october|november|december) \d\d?)", fwc_hearing_soup)
                    fwc_hearing_day_month = fwc_hearing_day_month.group()
                    if fwc_hearing_day_month is not None: 
                        fwc_hearing_date_string = str(fwc_hearing_day_month) + " " + str(fwc_hearing_year)
                              
                        hearing_date = get_date(fwc_hearing_date_string)
                        print("hearing date= " + str(hearing_date))

                #Submission Date
                if "final written submissions:" not in souptext_four: submission_date = "no submission"  
                else:
                    fwc_submission_soup = re.search("(?<=final written submissions:).*", souptext_four)
                    fwc_submission_soup = fwc_submission_soup.group()
                    #print(fwc_submission_soup)
                    
                    fwc_submission_day_month_applicant = re.search(
                        "(?<=applicant:)(\d\d? (january|february|march|april|may|june|july|august|september|october|november|december))|(\d\d? (january|february|march|april|may|june|july|august|september|october|november|december))", fwc_submission_soup)
                    
                    fwc_submission_day_month_applicant = fwc_submission_day_month_applicant.group()  
                    #print(fwc_submission_day_month_applicant)
                    
                    fwc_submission_day_month_respondent = re.search(
                        "(?<=respondent:)(\d\d? (january|february|march|april|may|june|july|august|september|october|november|december))|(\d\d? (january|february|march|april|may|june|july|august|september|october|november|december))", fwc_submission_soup)
                    
                    fwc_submission_day_month_respondent = fwc_submission_day_month_respondent.group()
                    print(fwc_submission_day_month_respondent)
                    
                    fwc_submission_year = re.search("\d\d\d\d", fwc_submission_soup)
                    fwc_submission_year = fwc_submission_year.group()
                    
                    fwc_submission_date_applicant = fwc_submission_day_month_applicant + " " + fwc_submission_year
                    fwc_submission_date_respondent = fwc_submission_day_month_respondent + " " + fwc_submission_year
                    #print(fwc_submission_date_applicant)

                    submission_date = get_date(fwc_submission_date_applicant)
                    #submission_date = the later of the two dates

            #Getting Time to Decide
            if type(hearing_date) is not str:
                time_to_decide_from_hearing = title_date - hearing_date
                time_to_decide_from_hearing = time_to_decide_from_hearing
                print(time_to_decide_from_hearing)
            else:
                time_to_decide_from_hearing = "N/A"
            
            if type(submission_date) is not str:
                print(submission_date)
                print(title_date)
                time_to_decide_from_last_submission = title_date - submission_date
                time_to_decide_from_last_submission = time_to_decide_from_last_submission
                print(time_to_decide_from_last_submission)
            else:
                time_to_decide_from_last_submission = "N/A"

    #CREATING DICT
            case_dict = {}

            case_dict[str(title)] = [

            judge_string, 
            str(case_matter), 
            bool(affirm_search) , 
            bool(allow_search) , 
            str(citation_list), 
            len(citation_list), 
            costs_order , 
            submission_date, 
            hearing_date, 
            title_date, 
            time_to_decide_from_last_submission, 
            time_to_decide_from_hearing, 
            registry, 
            division]
     
            print(case_dict)
            print(len(case_dict[str(title)]))

            lcdl = list(case_dictlist)

            #Edits to Dictlist below

            #Normalising Judge names for Dict
            for item in case_dict:
                case_dict[item][0] = strip_judge_name(case_dict[item][0])
            
            case_dictlist.append(dict(case_dict))

            #Normalsing '&' and 'and'
            
            
            
            #me testing out accessing dictlists
            n = n + 1
            print(case_dictlist[n][str(title)][0])
            print(case_dictlist[n])
            print(case_dictlist[n].values())
            
            #Exporting casedict to spreadsheet row
            excel_export(case_dict)
            wb.save(filename = dest_filename)

            #Readability barriers
            print("--------")
            f.write("------ \n")

        except Exception:
            #there is no case that can stop this loop; they are skipped.
            print("-"*60)
            traceback.print_exc(file=sys.stdout)
            print("-"*60)
            print("Exception caught: case skipped.")
            f.write(" ///// EXCEPTION CAUGHT ///// " + str(traceback.print_exc(file=sys.stdout)) + str(title) + "\n")
            continue


#SPREADSHEET FORMATTING         
#Spreadsheet sort/filter headings
ws1.auto_filter.ref = "B3:P3"
ws1.auto_filter.add_sort_condition("B4:P"+str(len(case_dictlist)))
wb.save(filename = dest_filename)

#Testing accessing the dictlist.
# #this works  
#print(case_dictlist[0])
#for nn in range(0, len(case_dictlist)):
#    print(case_dictlist[nn].values())
#    print(case_dictlist[nn].keys())


#this cycles through the case_dict values for the judge names. you MUST put .values() in a list wrapper because it returns
# a view and now a list. A view cant be indexed. Need to turn into a list to index
# e.g. [0][2] is accessing the tuple with 0 and a tuple value (judge etc) with 2
#SUCCES! Below is a list comprehension for your casedictlist
affirm_judge_collection = Counter([list(case_dictlist[x].values())[0][0] for x in range(0, len(case_dictlist)) if list(case_dictlist[x].values())[0][2] == True])
allowed_judge_collection = Counter([list(case_dictlist[x].values())[0][0] for x in range(0, len(case_dictlist)) if list(case_dictlist[x].values())[0][2] == False])
combined_judge_collection = allowed_judge_collection + affirm_judge_collection
judge_success_rates = get_success_rate(affirm_judge_collection,combined_judge_collection)

#IMPORTING CSV
#if mode == "import"
#with open('curated.csv', dialect='excel') as csvfile:
# for row in spamreader:
#  case_dictlist.append({row['Case Title']:[row['Judge'], row['Case Matter'], row['Plaintiff fail]_________]})
#  if row['Plaintiff success'] = True: allowed_citation_list = allowed_citation_list.append(row['Citations'])
#  elif row['Plaintiff fail'] = True: affirm_citation_list = affirm_citation_list.append(row['Citations'])

judge_weighted_rates = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[0]
    
judge_weighted_rate_mean = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[1]
    
judge_unweighted_rate_mean = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[2]

judge_unweighted_stdev = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[3]
    
judge_weighted_stdev = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[4]

#avg_judges = [dict(judge_unweighted_rate_mean).keys()]
#print(avg_judges)


#CASE CITATION OPERATIONS
#Creating the collections
affirm_citation_collection = aggregrate_success_fail(
    tuple(allowed_citation_list), tuple(affirm_citation_list))[0]

allowed_citation_collection = aggregrate_success_fail(
    tuple(allowed_citation_list), tuple(affirm_citation_list))[1]

combined_citation_collection = aggregrate_success_fail(
    tuple(allowed_citation_list), tuple(affirm_citation_list))[2] 

citation_success_rate = get_success_rate(
    affirm_citation_collection, combined_citation_collection)

#Getting weighting, weighted mean, unweighted mean, and unweighted stdev
#Getting weightings (e.g. rate * citation occurence)
citation_weighted_rates = mean_stdev_calculation(
    combined_citation_collection, affirm_citation_collection, citation_success_rate)[0]

#Getting mean of weighted citation success rates
citation_weighted_rate_mean = mean_stdev_calculation(
    combined_citation_collection, affirm_citation_collection, citation_success_rate)[1]

#Getting mean of unweighted citation success rates
citation_unweighted_rate_mean = mean_stdev_calculation(
    combined_citation_collection, affirm_citation_collection, citation_success_rate)[2]                                       

#Getting mean of unweighted citation standard deviation
citation_unweighted_stdev = mean_stdev_calculation(
    combined_citation_collection, affirm_citation_collection, citation_success_rate)[3]

#Getting mean of weighted citation standard deviation
citation_weighted_stdev = mean_stdev_calculation(
    combined_judge_collection, affirm_judge_collection, judge_success_rates)[3]

#Printing results
print("UNWEIGHTED PERCENTAGE ALLOWED PER JUDGE")
print(str(judge_success_rates))
print("SAMPLE SIZE")
print(combined_judge_collection)

#CITATION LENGTH AND SUCCESS CALCULATIONS
y_result = []
x_citation_length = []

for x in range(0, len(case_dictlist)):
    #appending all citation lengths
    x_citation_length.append((list(case_dictlist[x].values())[0][5]))
    #appending int(1 or -1) depending on value of affirm_search
    if list(case_dictlist[x].values())[0][2] == True:
        y_result.append(int(-1))
    elif list(case_dictlist[x].values())[0][2] == False:
        y_result.append(int(1))

print(x_citation_length)
print(y_result)

citationlength_allowed_crosscor = numpy.correlate(x_citation_length, y_result)
print("correlation cross-correlation   x: len(citations) , y: result")
print(citationlength_allowed_crosscor)

citationlength_allowed_polyfit = numpy.polyfit(x_citation_length, y_result, 2)
print("polyfit    x: len(citations) , y: result")
print(citationlength_allowed_polyfit)

citationlength_allowed_corrcoef= numpy.corrcoef(x_citation_length, y_result)
print("correlation coefficient x: cases cited    y: result)")
print(citationlength_allowed_corrcoef)

#CITATION LENGTH AND COSTS CALCULATIONS
#For FCA/FCCA only
citation_length_if_dollarcosts_list = []
dollar_costs_list = []

#if court == "FCCA" or "FCCA"

if court == "FCCA":
    for x in range( 0 , len(case_dictlist) ):
        #this is (1) checking the costs section to see if its a number and (2) checking if citations > 0
        #this is not working
        #list(case_dictlist[x].values())[0][6] == int and 
        if int(list(case_dictlist[x].values())[0][5]) > 0 and isinstance(list(case_dictlist[x].values())[0][6], int) == True:
                citation_length_if_dollarcosts_list.append(list(case_dictlist[x].values())[0][5]) #this is appending len(citations) (y value)
                dollar_costs_list.append(list(case_dictlist[x].values())[0][6]) ##this is appending costs order (x value)
        
    print(citation_length_if_dollarcosts_list)
    print(dollar_costs_list)

    citation_length_dollarcosts_corrcoef= numpy.corrcoef(citation_length_if_dollarcosts_list, dollar_costs_list)
    print("Citation count - Costs amount correlation")
    print(citation_length_dollarcosts_corrcoef)
else:
    citation_length_dollarcosts_corrcoef = "not applicable: no $Costs in %a".format(court)

#END OF FILE PRINT
#this will need to be openpyxl soon
#Writing the results file
f.write("Crawl conducted at " + str(datetime.datetime.now()) + "\n")
f.write("The case number range of this crawl was " + str(case_range) + " . This analysis was of the " + court + " for " + str(year_range) + "\n")
f.write("There were " + str(appeal_denied_count) + " confirmed denied appeals. There were " + str(appeal_allowed_count) + " confirmed allowed appeals \n")
rate = (100 * (appeal_allowed_count / (appeal_denied_count + appeal_allowed_count)))
f.write("confirmed appeal success rate = "+ str(rate) + " % \n")
f.write("The results of " + str(other_count) + " cases are unknown \n")
f.write("case meta dictionary")
f.write(str(case_dictlist))

#Writing the stats file
f_two.write("Crawl conducted at " + str(datetime.datetime.now()) + "\n")
f_two.write("Judges: unweighted appeal allowed %" + "\n" + str(judge_unweighted_rate_mean) + "\n")
f_two.write("Judges: unweighted stdev" + "\n" + str(judge_unweighted_stdev)+ "\n")
f_two.write("Judges % allowed dataset (unweighted) " + str(judge_success_rates) + "\n")
f_two.write("citation: unweighted appeal allowed %" + "\n" + str(citation_unweighted_rate_mean) + "\n")
f_two.write("citation: unweighted stdev" + "\n" + str(citation_unweighted_stdev) + "\n")
f_two.write("citation: weighted appeal allowed %" + "\n" + str(citation_weighted_rate_mean) + "\n")
f_two.write("citation: weighted stdev" + "\n" + str(citation_weighted_stdev) + "\n")
f_two.write("Citation % allowed dataset (unweighted) " + str(citation_success_rate)+ "\n")
f_two.write("correlation coefficient " + str(citationlength_allowed_corrcoef)+ "\n")
f_two.write("correlation cross-correlation   x: len(citations) , y: result " + str(citationlength_allowed_crosscor) + "\n")
f_two.write("Citation count - Costs amount correlation"+ "\n")
f_two.write(str(citation_length_dollarcosts_corrcoef)+ "\n")
f.close()
f_two.close()
